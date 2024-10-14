from simon import SimonCipher
import xlsxwriter
import openpyxl
import os
import random

def simon_encrypt(key, plaintext):
    cipher = SimonCipher(key)
    ciphertext = cipher.encrypt(plaintext)
    return ciphertext

def modify_bit(byte_string, bit_position):
    # Modify the bit at the specified position in the byte string
    byte_index = bit_position // 8
    bit_offset = bit_position % 8
    modified_byte = byte_string[byte_index] ^ (1 << bit_offset)
    return byte_string[:byte_index] + bytes([modified_byte]) + byte_string[byte_index + 1:]

def avalanche_effect(original_ciphertext, modified_ciphertext):
    if len(original_ciphertext) != len(modified_ciphertext):
        raise ValueError("Ciphertexts must be of equal length")

    num_bits = len(original_ciphertext) * 8
    diff_bits = sum(
        bin(original_ciphertext[i] ^ modified_ciphertext[i]).count('1')
        for i in range(len(original_ciphertext))
    )
    return diff_bits / num_bits * 100  # Calculate percentage of differing bits

def main():
    # Read the Excel file
    workbook = openpyxl.load_workbook('bank_1MB.xlsx')
    sheet = workbook.active

    # Convert the content of the Excel sheet to bytes
    plaintext = sheet.cell(row=1, column=1).value.encode('utf-8')

    # Generate a random SIMON key
    key_size = os.urandom(64)  # Choose appropriate key size (e.g., 64 for SIMON32/64)
    key = bytes([random.randint(0, 255) for _ in range(int(key_size))])

    # Encrypt the plaintext using SIMON
    original_ciphertext = simon_encrypt(key, plaintext)

    # Create a workbook to store results
    workbook = xlsxwriter.Workbook('simon_avalanche_effects.xlsx')
    worksheet = workbook.add_worksheet()

    # Header
    worksheet.write('A1', 'Bit Position')
    worksheet.write('B1', 'Avalanche Effect (%)')

    # Modify each bit in the ciphertext and calculate avalanche effect
    for bit_position in range(len(original_ciphertext) * 8):
        modified_ciphertext = modify_bit(original_ciphertext, bit_position)
        effect = avalanche_effect(original_ciphertext, modified_ciphertext)
        worksheet.write(bit_position + 1, 0, bit_position)
        worksheet.write(bit_position + 1, 1, effect)

    workbook.close()
    print("Avalanche effects written to 'simon_avalanche_effects.xlsx'.")

if __name__ == "__main__":
    main()
