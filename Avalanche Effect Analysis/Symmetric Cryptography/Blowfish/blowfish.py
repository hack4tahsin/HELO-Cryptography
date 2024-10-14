from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.backends import default_backend
import os
import xlsxwriter
import openpyxl
import random


def blowfish_encrypt(key, plaintext):
    iv = os.urandom(8)  # Initialization vector
    cipher = Cipher(algorithms.Blowfish(key), modes.CFB(iv), backend=default_backend())
    encryptor = cipher.encryptor()
    ciphertext = encryptor.update(plaintext) + encryptor.finalize()

    return ciphertext


def modify_bit(byte_string, bit_position):
    # Modify the bit at the specified position in the byte string
    byte_index = bit_position // 8
    bit_offset = bit_position % 8
    modified_byte = byte_string[byte_index] ^ (1 << bit_offset)
    return byte_string[:byte_index] + bytes([modified_byte]) + byte_string[byte_index + 1:]


def calculate_avalanche_effect(original_ciphertext, modified_ciphertext):
    if len(original_ciphertext) != len(modified_ciphertext):
        raise ValueError("Ciphertexts must be of equal length")

    num_bits = len(original_ciphertext) * 8
    diff_bits = sum(
        bin(original_ciphertext[i] ^ modified_ciphertext[i]).count('1') for i in range(len(original_ciphertext)))
    return diff_bits / num_bits * 100  # Calculate percentage of differing bits


def main():
    # Read the Excel file
    workbook = openpyxl.load_workbook('bank_100MB.xlsx')
    sheet = workbook.active

    # Convert the content of the Excel sheet to bytes
    plaintext = sheet.cell(row=1, column=1).value.encode('utf-8')

    # Generate a random AES key
    key = bytes([random.randint(0, 255) for _ in range(16)])

    # Encrypt the plaintext
    original_ciphertext = blowfish_encrypt(key, plaintext)

    # Modify each bit in the plaintext and calculate avalanche effect
    avalanche_effects = []
    for bit_position in range(len(plaintext) * 8):
        modified_plaintext = modify_bit(plaintext, bit_position)
        modified_ciphertext = blowfish_encrypt(key, modified_plaintext)
        avalanche_effect = calculate_avalanche_effect(original_ciphertext, modified_ciphertext)
        avalanche_effects.append(avalanche_effect)

    # Write the avalanche effects to a new Excel file
    output_workbook = xlsxwriter.Workbook('avalancheEffects.xlsx')
    output_sheet = output_workbook.add_worksheet()

    output_sheet.write('A1', 'Bit Position')
    output_sheet.write('B1', 'Avalanche Effect (%)')

    for i, avalanche_effect in enumerate(avalanche_effects):
        output_sheet.write(i + 1, 0, i)  # Bit Position
        output_sheet.write(i + 1, 1, avalanche_effect)  # Avalanche Effect (%)

    output_workbook.close()
    print("Avalanche effects written to 'avalanche_effects.xlsx'.")


if __name__ == "__main__":
    main()
