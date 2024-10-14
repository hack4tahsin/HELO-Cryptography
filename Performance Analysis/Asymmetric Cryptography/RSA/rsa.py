# # import time
# #
# # start = time.time()
# #
# # import base64
# # from cryptography.hazmat.backends import default_backend
# # from cryptography.hazmat.primitives.asymmetric import rsa
# # from cryptography.hazmat.primitives import serialization
# # from cryptography.hazmat.primitives import hashes
# # from cryptography.hazmat.primitives.asymmetric import padding
# # import bcrypt
# #
# # def utf8(s: bytes):
# #     return str(s, 'utf-8')
# #
# #
# # private_key = rsa.generate_private_key(
# #     public_exponent=65537,
# #     key_size=4096,
# #     backend=default_backend()
# # )
# # public_key = private_key.public_key()
# #
# #
# # private_pem = private_key.private_bytes(
# #     encoding=serialization.Encoding.PEM,
# #     format=serialization.PrivateFormat.PKCS8,
# #     encryption_algorithm=serialization.NoEncryption()
# # )
# #
# # private_derived_key = bcrypt.kdf(private_pem, salt=b'salt', desired_key_bytes=32, rounds=100)
# #
# # with open('private_key.pem', 'wb') as f:
# #     f.write(private_pem)
# #
# # public_pem = public_key.public_bytes(
# #     encoding=serialization.Encoding.PEM,
# #     format=serialization.PublicFormat.SubjectPublicKeyInfo
# # )
# #
# # public_derived_key = bcrypt.kdf(public_pem, salt=b'salt', desired_key_bytes=32, rounds=100)
# #
# # with open('public_key.pem', 'wb') as f:
# #     f.write(public_pem)
# #
# #
# # with open("private_key.pem", "rb") as key_file:
# #     private_key = serialization.load_pem_private_key(
# #         key_file.read(),
# #         password=None,
# #         backend=default_backend()
# #     )
# #
# # with open("public_key.pem", "rb") as key_file:
# #     public_key = serialization.load_pem_public_key(
# #         key_file.read(),
# #         backend=default_backend()
# #     )
# #
# #
# # plaintext = b'this is the correct plaintext!'
# # print(f'plaintext: \033[1;33m{utf8(plaintext)}\033[0m')
# # encrypted = base64.b64encode(public_key.encrypt(
# #     plaintext,
# #     padding.OAEP(
# #         mgf=padding.MGF1(algorithm=hashes.SHA256()),
# #         algorithm=hashes.SHA256(),
# #         label=None
# #     )
# # ))
# # print(f'encrypted: \033[1;32m{utf8(encrypted)}\033[0m')
# #
# #
# # decrypted = private_key.decrypt(
# #     base64.b64decode(encrypted),
# #     padding.OAEP(
# #         mgf=padding.MGF1(algorithm=hashes.SHA256()),
# #         algorithm=hashes.SHA256(),
# #         label=None
# #     )
# # )
# # print(f'decrypted: \033[1;31m{utf8(decrypted)}\033[0m')
# #
# #
# # end = time.time()
# #
# # total_time = end - start
# #
# # print("Total execution time of RSA: {:.2f} s".format(total_time))
#
#
# import os
# from cryptography.hazmat.primitives.asymmetric import rsa, padding
# from cryptography.hazmat.primitives import serialization
# from cryptography.hazmat.backends import default_backend
#
#
# def generate_rsa_key_pair():
#     # Generate RSA key pair
#     private_key = rsa.generate_private_key(
#         public_exponent=65537,
#         key_size=2048,
#         backend=default_backend()
#     )
#     public_key = private_key.public_key()
#     return private_key, public_key
#
#
# def save_key_to_file(key, filename):
#     # Save key to a file
#     with open(filename, 'wb') as f:
#         f.write(key)
#
#
# def load_key_from_file(filename):
#     # Load key from a file
#     with open(filename, 'rb') as f:
#         key = f.read()
#     return key
#
#
# def encrypt_file(file_path, public_key):
#     # Encrypt the file using RSA public key
#     with open(file_path, 'rb') as f:
#         data = f.read()
#
#     public_key_obj = serialization.load_pem_public_key(public_key, backend=default_backend())
#     encrypted_data = public_key_obj.encrypt(
#         data,
#         padding.OAEP(
#             mgf=padding.MGF1(algorithm=padding.ALGORITHMS.SHA256),
#             algorithm=padding.ALGORITHMS.SHA256,
#             label=None
#         )
#     )
#
#     with open(file_path + '.enc', 'wb') as f:
#         f.write(encrypted_data)
#     os.remove(file_path)
#
#
# def decrypt_file(file_path, private_key):
#     # Decrypt the file using RSA private key
#     with open(file_path, 'rb') as f:
#         encrypted_data = f.read()
#
#     private_key_obj = serialization.load_pem_private_key(private_key, password=None, backend=default_backend())
#     decrypted_data = private_key_obj.decrypt(
#         encrypted_data,
#         padding.OAEP(
#             mgf=padding.MGF1(algorithm=padding.ALGORITHMS.SHA256),
#             algorithm=padding.ALGORITHMS.SHA256,
#             label=None
#         )
#     )
#
#     with open(file_path[:-4], 'wb') as f:
#         f.write(decrypted_data)
#     os.remove(file_path)
#
#
# # Example usage
# # Generate RSA key pair
# private_key, public_key = generate_rsa_key_pair()
#
# # Save the keys to files
# save_key_to_file(private_key.private_bytes(encoding=serialization.Encoding.PEM,
#                                            format=serialization.PrivateFormat.TraditionalOpenSSL,
#                                            encryption_algorithm=serialization.NoEncryption()), 'private_key.pem')
# save_key_to_file(public_key.public_bytes(encoding=serialization.Encoding.PEM,
#                                          format=serialization.PublicFormat.SubjectPublicKeyInfo), 'public_key.pem')
#
# # Encrypting the file
# encrypt_file("bank_1MB.xlsx", public_key)
#
# # Decrypting the file
# decrypt_file("example.xlsx.enc", private_key.private_bytes(encoding=serialization.Encoding.PEM,
#                                                           format=serialization.PrivateFormat.TraditionalOpenSSL,
#                                                           encryption_algorithm=serialization.NoEncryption()))


# from cryptography.hazmat.primitives import serialization, hashes
# from cryptography.hazmat.primitives.asymmetric import rsa
# from cryptography.hazmat.primitives.asymmetric import padding
# from cryptography.hazmat.primitives import padding as symmetric_padding
# from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
# import os
# import openpyxl
# import bcrypt
# import time
#
# # Generate RSA key pair for asymmetric encryption
# def generate_key_pair():
#     private_key = rsa.generate_private_key(
#         public_exponent=65537,
#         key_size=1024
#     )
#     public_key = private_key.public_key()
#     return private_key, public_key
#
#
# # Encrypt data using RSA public key
# def rsa_encrypt(data, public_key):
#     encrypted_data = public_key.encrypt(
#         data,
#         padding.OAEP(
#             mgf=padding.MGF1(algorithm=hashes.SHA256()),
#             algorithm=hashes.SHA256(),
#             label=None
#         )
#     )
#     return encrypted_data
#
#
# # Decrypt data using RSA private key
# def rsa_decrypt(encrypted_data, private_key):
#     decrypted_data = private_key.decrypt(
#         encrypted_data,
#         padding.OAEP(
#             mgf=padding.MGF1(algorithm=hashes.SHA256()),
#             algorithm=hashes.SHA256(),
#             label=None
#         )
#     )
#
#     return decrypted_data
#
# # Encrypt XLSX file using AES
# def encrypt_xlsx(file_path, symmetric_key):
#     workbook = openpyxl.load_workbook(file_path)
#     for sheet in workbook:
#         for row in sheet.iter_rows():
#             for cell in row:
#                 if cell.value:
#                     cipher = Cipher(algorithms.AES(symmetric_key), modes.ECB())
#                     encryptor = cipher.encryptor()
#                     padder = symmetric_padding.PKCS7(algorithms.AES.block_size).padder()
#                     padded_data = padder.update(str(cell.value).encode()) + padder.finalize()
#                     encrypted_data = encryptor.update(padded_data) + encryptor.finalize()
#                     cell.value = encrypted_data.hex()
#     workbook.save(file_path)
#
#
# # Decrypt XLSX file using AES
# def decrypt_xlsx(file_path, symmetric_key):
#     workbook = openpyxl.load_workbook(file_path)
#     for sheet in workbook:
#         for row in sheet.iter_rows():
#             for cell in row:
#                 if cell.value:
#                     cipher = Cipher(algorithms.AES(symmetric_key), modes.ECB())
#                     decryptor = cipher.decryptor()
#                     decrypted_data = decryptor.update(bytes.fromhex(cell.value)) + decryptor.finalize()
#                     unpadder = symmetric_padding.PKCS7(algorithms.AES.block_size).unpadder()
#                     unpadded_data = unpadder.update(decrypted_data) + unpadder.finalize()
#                     cell.value = unpadded_data.decode()
#     workbook.save(file_path)
#
#
# # Main function for encryption and decryption
# def main():
#     start_1 = time.time()
#
#     # Generate RSA key pair
#     private_key, public_key = generate_key_pair()
#
#     # Encrypt XLSX file
#     file_path = "bank_1MB.xlsx"
#
#     end_1 = time.time()
#     total_1 = end_1 - start_1
#
#     symmetric_key = os.urandom(32)  # Generate a random 256-bit symmetric key
#     encrypt_xlsx(file_path, symmetric_key)
#
#     start_2 = time.time()
#
#     # Encrypt the symmetric key using RSA public key
#     encrypted_symmetric_key = rsa_encrypt(symmetric_key, public_key)
#
#     # Save encrypted symmetric key to a file
#     with open("encrypted_symmetric_key.pem", "wb") as key_file:
#         key_file.write(encrypted_symmetric_key)
#
#     # Decrypt the symmetric key using RSA private key
#     with open("encrypted_symmetric_key.pem", "rb") as key_file:
#         encrypted_symmetric_key = key_file.read()
#     decrypted_symmetric_key = rsa_decrypt(encrypted_symmetric_key, private_key)
#
#     end_2 = time.time()
#
#     total_2 = end_2 - start_2
#
#     # Decrypt XLSX file
#     decrypt_xlsx(file_path, decrypted_symmetric_key)
#
#     total_time = total_1 + total_2
#
#     print("Total execution time of RSA: {:.2f} s".format(total_time))
#
#
# if __name__ == "__main__":
#     main()


import base64
import bcrypt
import openpyxl
from cryptography.hazmat.backends import default_backend
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.asymmetric import padding
from cryptography.hazmat.primitives.asymmetric import rsa
from cryptography.hazmat.primitives import serialization
import time


start = time.time()

# Generate RSA keys for asymmetric encryption
def generate_rsa_keys():
    private_key = rsa.generate_private_key(
        public_exponent=65537,
        key_size=1024,
        backend=default_backend()
    )
    public_key = private_key.public_key()
    return private_key, public_key


# Save RSA private key to file
def save_private_key(private_key, filename, password):
    pem = private_key.private_bytes(
        encoding=serialization.Encoding.PEM,
        format=serialization.PrivateFormat.PKCS8,
        encryption_algorithm=serialization.BestAvailableEncryption(password.encode())
    )
    with open(filename, 'wb') as f:
        f.write(pem)


# Save RSA public key to file
def save_public_key(public_key, filename):
    pem = public_key.public_bytes(
        encoding=serialization.Encoding.PEM,
        format=serialization.PublicFormat.SubjectPublicKeyInfo
    )
    with open(filename, 'wb') as f:
        f.write(pem)


# Load RSA private key from file
def load_private_key(filename, password):
    with open(filename, 'rb') as f:
        pem_data = f.read()
    private_key = serialization.load_pem_private_key(
        pem_data,
        password=password.encode(),
        backend=default_backend()
    )
    return private_key


# Encrypt data using RSA public key
def rsa_encrypt(data, public_key):
    ciphertext = public_key.encrypt(
        data,
        padding.OAEP(
            mgf=padding.MGF1(algorithm=hashes.SHA256()),
            algorithm=hashes.SHA256(),
            label=None
        )
    )
    return ciphertext


# Decrypt data using RSA private key
def rsa_decrypt(ciphertext, private_key):
    plaintext = private_key.decrypt(
        ciphertext,
        padding.OAEP(
            mgf=padding.MGF1(algorithm=hashes.SHA256()),
            algorithm=hashes.SHA256(),
            label=None
        )
    )
    return plaintext


# # Hash password using bcrypt
# def hash_password(password):
#     hashed = bcrypt.hashpw(password.encode(), bcrypt.gensalt())
#     return hashed
#
#
# # Verify password against hash using bcrypt
# def verify_password(password, hashed_password):
#     return bcrypt.checkpw(password.encode(), hashed_password)


# Encrypt Excel file using RSA public key
def encrypt_excel(file_path, public_key_file):
    workbook = openpyxl.load_workbook(file_path)

    # Load public key
    with open(public_key_file, 'rb') as f:
        public_key = serialization.load_pem_public_key(
            f.read(),
            backend=default_backend()
        )

    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        data = str(ws.values).encode()
        encrypted_data = rsa_encrypt(data, public_key)
        ws['A1'] = base64.b64encode(encrypted_data).decode()

    workbook.save(file_path)


# Decrypt Excel file using RSA private key
def decrypt_excel(file_path, private_key_file, password):
    # Load private key
    private_key = load_private_key(private_key_file, password)

    workbook = openpyxl.load_workbook(file_path)

    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        encrypted_data = base64.b64decode(ws['A1'].value)
        decrypted_data = rsa_decrypt(encrypted_data, private_key)
        ws['A1'] = eval(decrypted_data.decode())

    workbook.save(file_path)


# Example usage
if __name__ == "__main__":
    # Generate RSA keys
    private_key, public_key = generate_rsa_keys()

    # Save private key (with password)
    save_private_key(private_key, 'private_key.pem', 'my_password')

    # Save public key
    save_public_key(public_key, 'public_key.pem')

    # Hash a password
    # hashed_password = hash_password('my_password')

    # Verify password
    # print(verify_password('my_password', hashed_password))  # True or False

    # Encrypt Excel file using RSA public key
    encrypt_excel('bank_1MB.xlsx', 'public_key.pem')
    #
    # # Decrypt Excel file using RSA private key (with password)
    # decrypt_excel('bank_1MB.xlsx', 'private_key.pem', 'my_password')

end = time.time()

total_time = end - start

print("Total execution time of RSA: {:.2f} s".format(total_time))