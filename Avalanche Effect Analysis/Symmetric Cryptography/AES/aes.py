import os

import openpyxl
import random
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.backends import default_backend
from cryptography.hazmat.primitives import padding

def aes_encrypt(key, plaintext):
    backend = default_backend()
    iv = b'\x00' * 16  # Initialization vector
    cipher = Cipher(algorithms.AES(key), modes.CBC(iv), backend=backend)
    encryptor = cipher.encryptor()
    padder = padding.PKCS7(algorithms.AES.block_size).padder()
    padded_data = padder.update(plaintext) + padder.finalize()
    ciphertext = encryptor.update(padded_data) + encryptor.finalize()
    return ciphertext

def avalanche_effect(original_ciphertext, modified_ciphertext):
    if len(original_ciphertext) != len(modified_ciphertext):
        raise ValueError("Ciphertexts must be of equal length")

    num_bits = len(original_ciphertext) * 8
    diff_bits = sum(
        bin(original_ciphertext[i] ^ modified_ciphertext[i]).count('1') for i in range(len(original_ciphertext)))
    return diff_bits / num_bits * 100  # Calculate percentage of differing bits

def modify_bit(byte_string, bit_position):
    # Modify the bit at the specified position in the byte string
    byte_index = bit_position // 8
    bit_offset = bit_position % 8
    modified_byte = byte_string[byte_index] ^ (1 << bit_offset)
    return byte_string[:byte_index] + bytes([modified_byte]) + byte_string[byte_index + 1:]

def main():
    # Create a workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Avalanche Effect"

    # Header row
    sheet['A1'] = "Bit Position"
    sheet['B1'] = "Avalanche Effect (%)"

    # Generate a random AES key
    # key = bytes([random.randint(0, 255) for _ in range(16)])

    key = os.urandom(16)

    file_path = '(1 MB) SampleSuperstore.csv'
    with open(file_path, 'rb') as file:
        plaintext = file.read()

    # Encrypt the plaintext
    original_ciphertext = aes_encrypt(key, plaintext)

    # Calculate avalanche effect for bit positions 0 to 20
    for bit_position in range(21):  # from 0 to 20 inclusive
        modified_plaintext = modify_bit(plaintext, bit_position)
        modified_ciphertext = aes_encrypt(key, modified_plaintext)
        effect = avalanche_effect(original_ciphertext, modified_ciphertext)

        # Save results to Excel sheet
        row = bit_position + 2  # starting from row 2 onwards
        sheet[f'A{row}'] = bit_position
        sheet[f'B{row}'] = effect

    # Save workbook to file
    workbook.save('avalanche_effects.xlsx')
    print("Results saved to avalanche_effects.xlsx")

if __name__ == "__main__":
    main()








# from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
# from cryptography.hazmat.backends import default_backend
# from cryptography.hazmat.primitives import padding
# import xlsxwriter
# import openpyxl
# import random
#
#
# def aes_encrypt(key, plaintext):
#     backend = default_backend()
#     iv = b'\x00' * 16  # Initialization vector
#     cipher = Cipher(algorithms.AES(key), modes.CBC(iv), backend=backend)
#     encryptor = cipher.encryptor()
#     padder = padding.PKCS7(algorithms.AES.block_size).padder()
#     padded_data = padder.update(plaintext) + padder.finalize()
#     ciphertext = encryptor.update(padded_data) + encryptor.finalize()
#     return ciphertext
#
#
# def modify_bit(byte_string, bit_position):
#     # Modify the bit at the specified position in the byte string
#     byte_index = bit_position // 8
#     bit_offset = bit_position % 8
#     modified_byte = byte_string[byte_index] ^ (1 << bit_offset)
#     return byte_string[:byte_index] + bytes([modified_byte]) + byte_string[byte_index + 1:]
#
#
# def calculate_avalanche_effect(original_ciphertext, modified_ciphertext):
#     if len(original_ciphertext) != len(modified_ciphertext):
#         raise ValueError("Ciphertexts must be of equal length")
#
#     num_bits = len(original_ciphertext) * 8
#     diff_bits = sum(
#         bin(original_ciphertext[i] ^ modified_ciphertext[i]).count('1') for i in range(len(original_ciphertext)))
#     return diff_bits / num_bits * 100  # Calculate percentage of differing bits
#
#
# def main():
#     # Read the Excel file
#     workbook = openpyxl.load_workbook('bank_1MB.xlsx')
#     sheet = workbook.active
#
#     # Convert the content of the Excel sheet to bytes
#     plaintext = sheet.cell(row=1, column=1).value.encode('utf-8')
#
#     # Generate a random AES key
#     key = bytes([random.randint(0, 255) for _ in range(16)])
#
#     # Encrypt the plaintext
#     original_ciphertext = aes_encrypt(key, plaintext)
#
#     # Modify each bit in the plaintext and calculate avalanche effect
#     avalanche_effects = []
#     for bit_position in range(len(plaintext) * 8):
#         modified_plaintext = modify_bit(plaintext, bit_position)
#         modified_ciphertext = aes_encrypt(key, modified_plaintext)
#         avalanche_effect = calculate_avalanche_effect(original_ciphertext, modified_ciphertext)
#         avalanche_effects.append(avalanche_effect)
#
#     # Write the avalanche effects to a new Excel file
#     output_workbook = xlsxwriter.Workbook('avalancheEffects.xlsx')
#     output_sheet = output_workbook.add_worksheet()
#
#     output_sheet.write('A1', 'Bit Position')
#     output_sheet.write('B1', 'Avalanche Effect (%)')
#
#     for i, avalanche_effect in enumerate(avalanche_effects):
#         output_sheet.write(i + 1, 0, i)  # Bit Position
#         output_sheet.write(i + 1, 1, avalanche_effect)  # Avalanche Effect (%)
#
#     output_workbook.close()
#     print("Avalanche effects written to 'avalanche_effects.xlsx'.")
#
#
# if __name__ == "__main__":
#     main()


# import os
# import time
# import bcrypt
# import psutil
# from cryptography.hazmat.backends import default_backend
# from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
# from cryptography.hazmat.primitives import padding as sym_padding
# from openpyxl import Workbook
# import bitstring
#
#
# # Function to generate a random AES key
# def generate_aes_key():
#     return os.urandom(32)  # 256-bit AES key
#
#
# # Function to hash AES key with bcrypt
# def hash_aes_key_with_bcrypt(aes_key):
#     return bcrypt.hashpw(aes_key, bcrypt.gensalt())
#
#
# # Function to verify AES key hashed with bcrypt
# def verify_aes_key_with_bcrypt(aes_key, bcrypt_hash):
#     return bcrypt.checkpw(aes_key, bcrypt_hash)
#
#
# # Function to encrypt a file using AES
# def encrypt_file(file_path, aes_key, output_path):
#     if not os.path.isfile(file_path):
#         raise FileNotFoundError(f"Input file not found: {file_path}")
#
#     with open(file_path, 'rb') as file:
#         data = file.read()
#
#     iv = os.urandom(16)  # Initialization vector for AES
#
#     padder = sym_padding.PKCS7(algorithms.AES.block_size).padder()
#     padded_data = padder.update(data) + padder.finalize()
#
#     cipher = Cipher(algorithms.AES(aes_key), modes.CFB(iv), backend=default_backend())
#     encryptor = cipher.encryptor()
#     encrypted_data = encryptor.update(padded_data) + encryptor.finalize()
#
#     output_dir = os.path.dirname(output_path)
#     if output_dir:  # Check if output_dir is not empty
#         os.makedirs(output_dir, exist_ok=True)
#
#     with open(output_path, 'wb') as file:
#         file.write(iv)
#         file.write(encrypted_data)
#
#     return iv, encrypted_data
#
#
# # Function to decrypt a file using AES
# def decrypt_file(encrypted_file_path, aes_key, output_path):
#     if not os.path.isfile(encrypted_file_path):
#         raise FileNotFoundError(f"Encrypted file not found: {encrypted_file_path}")
#
#     with open(encrypted_file_path, 'rb') as file:
#         iv = file.read(16)
#         encrypted_data = file.read()
#
#     cipher = Cipher(algorithms.AES(aes_key), modes.CFB(iv), backend=default_backend())
#     decryptor = cipher.decryptor()
#     decrypted_padded_data = decryptor.update(encrypted_data) + decryptor.finalize()
#
#     unpadder = sym_padding.PKCS7(algorithms.AES.block_size).unpadder()
#     data = unpadder.update(decrypted_padded_data) + unpadder.finalize()
#
#     output_dir = os.path.dirname(output_path)
#     if output_dir:  # Check if output_dir is not empty
#         os.makedirs(output_dir, exist_ok=True)
#
#     with open(output_path, 'wb') as file:
#         file.write(data)
#
#     return data
#
#
# # Function to measure performance
# def measure_performance():
#     process = psutil.Process(os.getpid())
#     cpu_times = process.cpu_times()
#     memory_info = process.memory_info()
#
#     cpu_usage = cpu_times.user
#     memory_usage = memory_info.rss / (1024 * 1024)
#
#     return cpu_usage, memory_usage
#
#
# # Function to flip a bit in the AES key
# def flip_bit(aes_key, bit_position):
#     bit_array = bitstring.BitArray(bytes=aes_key)
#     bit_array.invert(bit_position)
#     return bit_array.tobytes()
#
#
# # Function to calculate the avalanche effect
# def calculate_avalanche_effect(original_data, modified_data):
#     diff_bits = sum(bin(x ^ y).count('1') for x, y in zip(original_data, modified_data))
#     total_bits = len(original_data) * 8
#     return (diff_bits / total_bits) * 100
#
#
# # Main function
# def main():
#     # Generate AES key
#     aes_key = generate_aes_key()
#
#     # Define file paths
#     input_file_path = '(10 MB) players_16.csv'
#     encrypted_file_path = 'output/encrypted_file.enc'
#     modified_encrypted_file_path = 'output/modified_encrypted_file.enc'
#
#     # Check if input file exists
#     if not os.path.isfile(input_file_path):
#         print(f"Input file not found: {input_file_path}")
#         return
#
#     # Encrypt file with original AES key
#     iv, encrypted_data = encrypt_file(input_file_path, aes_key, encrypted_file_path)
#
#     # Create an Excel workbook to store the avalanche effect results
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Avalanche Effect Results"
#     ws.append(["Bit Position", "Avalanche Effect (%)"])
#
#     for bit_position in range(21):
#         # Flip a bit in the AES key
#         modified_aes_key = flip_bit(aes_key, bit_position)
#
#         # Encrypt the file with the modified AES key
#         _, modified_encrypted_data = encrypt_file(input_file_path, modified_aes_key, modified_encrypted_file_path)
#
#         # Calculate the avalanche effect
#         avalanche_effect = calculate_avalanche_effect(encrypted_data, modified_encrypted_data)
#
#         # Write the results to the Excel file
#         ws.append([bit_position, avalanche_effect])
#
#     # Save the Excel file
#     output_excel_path = "output/avalanche_effect_results.xlsx"
#     os.makedirs(os.path.dirname(output_excel_path), exist_ok=True)
#     wb.save(output_excel_path)
#     print(f"Avalanche effect results saved to {output_excel_path}")
#
#
# if __name__ == "__main__":
#     main()
